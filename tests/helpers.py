"""Shared test helpers for xlsxturbo integration tests."""

from __future__ import annotations

import os
import tempfile
from pathlib import Path
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from openpyxl import load_workbook
    from openpyxl.workbook.workbook import Workbook
    from openpyxl.worksheet.worksheet import Worksheet

    HAS_OPENPYXL = True
else:
    try:
        from openpyxl import load_workbook

        HAS_OPENPYXL = True
    except ImportError:
        load_workbook = None
        HAS_OPENPYXL = False

__all__ = [
    "HAS_OPENPYXL",
    "REPO_ROOT",
    "TIMESTAMPED_PART",
    "TINY_PNG_B64",
    "active_ws",
    "get_temp_path",
    "load_workbook",
    "repo_checkout_available",
]

# Where the repository root would be, if these tests are running from a checkout.
REPO_ROOT = Path(__file__).resolve().parent.parent

# The one archive member that legitimately differs between two runs: it records
# the creation time at one-second resolution. Any test comparing two exports
# byte for byte has to exclude it, or it fails whenever the two writes land on
# opposite sides of a second boundary. `tests/test_stability_policy.py` owns the
# assertion that this is the ONLY such part.
TIMESTAMPED_PART = "docProps/core.xml"


def repo_checkout_available() -> bool:
    """Whether these tests can read repository files, not just the package.

    Most of this suite exercises the installed ``xlsxturbo`` package and works
    anywhere. A few modules instead audit *repository* artifacts -- the MkDocs
    config, the generated capability matrix, the generator script -- and those
    files exist only in a source checkout.

    The release workflow's smoke-test job copies **only** ``tests/`` next to an
    installed wheel and runs pytest from there, deliberately, so that pytest's
    rootdir handling cannot pull in the local ``python/`` tree instead of the
    wheel. Repository-auditing modules therefore have nothing to read and must
    skip rather than fail: they are not testing the wheel, which is the only
    thing that job exists to test.

    This is not hypothetical. Both such modules were added after v0.18.0, so the
    v0.19.0 release was the first to run them in that job, and all 16 of their
    tests failed there while passing locally.

    Returns:
        True when ``pyproject.toml`` sits above ``tests/``, which is the
        definitive marker of the source tree. Inside a checkout, a *missing*
        `mkdocs.yml` or capability matrix stays a real failure -- this guard
        distinguishes "no repository here" from "the repository is broken".
    """
    return (REPO_ROOT / "pyproject.toml").is_file()

# Base64 encoding of the smallest valid PNG: a single 1x1 white pixel.
TINY_PNG_B64 = "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR42mP8z8BQDwAEhQGAhKmMIQAAAABJRU5ErkJggg=="


def get_temp_path(suffix: str = ".xlsx") -> str:
    """Return a temporary file path with its handle closed.

    The handle is closed immediately so Windows allows the file to be
    reopened and rewritten by the library under test.

    Args:
        suffix: File extension (including the dot) for the temp path.

    Returns:
        The path to a newly created, empty temporary file.
    """
    fd, path = tempfile.mkstemp(suffix=suffix)
    os.close(fd)
    return path


def active_ws(wb: Workbook) -> Worksheet:
    """Return the active worksheet of ``wb``, asserting one exists.

    openpyxl types :attr:`Workbook.active` as ``Worksheet | None``; in these
    tests a freshly written workbook always has an active sheet, so this helper
    narrows the type for callers.
    """
    ws = wb.active
    assert ws is not None, "workbook has no active worksheet"
    return ws
