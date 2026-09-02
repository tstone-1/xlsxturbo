"""End-to-end boundary tests: where a written value changes kind.

The Rust suite pins these boundaries at the functions that decide them
(``src/write.rs``, ``src/parse/boundaries.rs``). This module checks the same
edges through the whole pipeline and reads the result back with a parser that
did not write it, because a boundary that holds in the decision function and is
lost two layers later is still a wrong file.

Two families, both chosen because crossing the line silently changes data
rather than raising:

- **Integers near 2^53.** Excel stores every number as an IEEE-754 double, so
  an integer past 2^53 cannot be written as a number without changing value.
  xlsxturbo writes those as text instead. The interesting cases are the two
  values either side of the cutoff, and ``i64::MIN``, whose magnitude cannot be
  taken with ``abs()`` at all.
- **Dates around Excel's 1900 leap-year bug.** Excel numbers dates as though
  1900-02-29 existed. Anything below serial 61 is therefore off by one against
  a real calendar, and xlsxturbo writes those as text rather than as a date
  that renders on the wrong day.
"""

from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING, Any

import pandas as pd
import pytest
import xlsxturbo

from tests.helpers import active_ws, load_workbook

if TYPE_CHECKING:
    from collections.abc import Callable

# The largest magnitude an IEEE-754 double represents exactly.
MAX_SAFE = 2**53

# Excel's first correctly numbered day, and the serial it carries.
FIRST_SANE_DATE = "1900-03-01"
FIRST_SANE_SERIAL = 61


def _cell(path: str, ref: str) -> Any:
    """Read one cell back from a written workbook."""
    workbook = load_workbook(path)
    try:
        return active_ws(workbook)[ref].value
    finally:
        workbook.close()


class TestIntegerPrecisionBoundary:
    """The 2^53 cutoff, from Python values through to the written cell."""

    @pytest.mark.parametrize(
        ("value", "written_as_number"),
        [
            (0, True),
            (MAX_SAFE - 1, True),
            (MAX_SAFE, True),  # exactly representable: a power of two
            (MAX_SAFE + 1, False),  # the first integer a double cannot hold
            (MAX_SAFE + 2, False),
            (-(MAX_SAFE - 1), True),
            (-MAX_SAFE, True),
            (-(MAX_SAFE + 1), False),
            (2**63 - 1, False),  # i64::MAX
            (-(2**63), False),  # i64::MIN -- its magnitude overflows abs()
            (2**64, False),  # past i64 entirely, so past the i64 fast path
            (-(2**70), False),
        ],
    )
    def test_the_cutoff_is_at_two_to_the_53(
        self, tmp_xlsx: str, value: int, written_as_number: bool
    ) -> None:
        """Values up to 2^53 are numbers; past it they become text."""
        frame = pd.DataFrame({"v": [value]}, dtype=object)
        xlsxturbo.df_to_xlsx(frame, tmp_xlsx)
        cell = _cell(tmp_xlsx, "A2")
        if written_as_number:
            assert cell == value
            assert not isinstance(cell, str)
        else:
            # Text, and the exact digits -- the point of the fallback is that
            # no precision is lost, so a stringified approximation would be a
            # failure dressed as a pass.
            assert cell == str(value)

    def test_the_string_fallback_keeps_every_digit(self, tmp_xlsx: str) -> None:
        """A value past the cutoff survives with all of its digits intact.

        Writing it as a number would round it to an even neighbour, so the
        assertion that matters is inequality with that neighbour, not merely
        that something string-shaped came back.
        """
        value = MAX_SAFE + 1
        frame = pd.DataFrame({"v": [value]}, dtype=object)
        xlsxturbo.df_to_xlsx(frame, tmp_xlsx)
        assert _cell(tmp_xlsx, "A2") == str(value)
        assert int(_cell(tmp_xlsx, "A2")) != int(float(value))

    def test_the_cutoff_applies_to_explicit_cells_too(self, tmp_xlsx: str) -> None:
        """The ``cells`` option goes through the same write path."""
        frame = pd.DataFrame({"v": [1]})
        xlsxturbo.df_to_xlsx(
            frame,
            tmp_xlsx,
            cells={"C1": MAX_SAFE, "C2": MAX_SAFE + 1},
        )
        assert _cell(tmp_xlsx, "C1") == MAX_SAFE
        assert _cell(tmp_xlsx, "C2") == str(MAX_SAFE + 1)

    def test_the_cutoff_applies_on_the_csv_path(
        self, tmp_xlsx_factory: Callable[..., str]
    ) -> None:
        """CSV values reach the cutoff through type detection, not PyO3.

        A separate code path -- ``CellValue::Integer`` rather than the Python
        integer extractor -- so it needs its own case.
        """
        csv_path = tmp_xlsx_factory(".csv")
        xlsx_path = tmp_xlsx_factory()
        Path(csv_path).write_text(
            f"safe,unsafe\n{MAX_SAFE},{MAX_SAFE + 1}\n", encoding="utf-8"
        )
        xlsxturbo.csv_to_xlsx(csv_path, xlsx_path)
        assert _cell(xlsx_path, "A2") == MAX_SAFE
        assert _cell(xlsx_path, "B2") == str(MAX_SAFE + 1)

    @pytest.mark.parametrize(
        "digits",
        [
            "12345678901234567890",  # 20 digits, the ordinary long identifier
            "9223372036854775808",  # i64::MAX + 1
            "9223372036854775807",  # i64::MAX: a control, see below
            "-9223372036854775808",  # i64::MIN: a control, see below
            "-9223372036854775809",  # i64::MIN - 1
            "18446744073709551615",  # u64::MAX
            "18446744073709551616",  # u64::MAX + 1, past every Rust integer
            "1" * 400,  # long enough that f64 parses it as infinity
        ],
    )
    def test_csv_integers_beyond_i64_are_written_as_text(
        self, tmp_xlsx_factory: Callable[..., str], digits: str
    ) -> None:
        """A CSV integer too large for i64 keeps every digit.

        The CSV path detects types by parsing text, and its integer attempt is
        an ``i64``. Until 1.3.1 a failed ``i64`` fell straight through to
        ``f64``, which parses any run of digits: ``12345678901234567890``
        reached the workbook as 12345678901234567000, and a 400-digit run
        parsed as infinity and reached it as an empty cell. Both contradict the
        guarantee ``test_the_cutoff_is_at_two_to_the_53`` holds on the
        DataFrame path.

        The assertion is a literal digit comparison, not a tolerance: a
        rounded value is still greater than 2^53, so any inequality against a
        bound would pass on the broken output.

        Two of the cases are controls rather than guards. ``i64::MAX`` and
        ``i64::MIN`` are inside the integer branch, so they were already text
        and stay green when the new screens are mutated out -- they are here to
        say where the boundary is. Measured: mutating both screens away reddens
        six of the eight.
        """
        csv_path = tmp_xlsx_factory(".csv")
        xlsx_path = tmp_xlsx_factory()
        Path(csv_path).write_text(f"v\n{digits}\n", encoding="utf-8")
        xlsxturbo.csv_to_xlsx(csv_path, xlsx_path)
        assert _cell(xlsx_path, "A2") == digits

    @pytest.mark.parametrize(
        ("text", "expected"),
        [
            ("9007199254740992", 2**53),  # the cutoff itself stays a number
            ("007", 7),  # leading zeros still normalise
            ("+42", 42),
            ("1e3", 1000.0),  # exponent notation stays on the float branch
            ("1.5", 1.5),
        ],
    )
    def test_the_csv_integer_screen_leaves_ordinary_numbers_alone(
        self, tmp_xlsx_factory: Callable[..., str], text: str, expected: object
    ) -> None:
        """The control for the screen above.

        Every case here was measured through the shipped 1.3.0 wheel before the
        screen was added and must come back unchanged, so a screen that widened
        into the float branch or into the i64 range fails here rather than
        silently turning working numbers into text.
        """
        csv_path = tmp_xlsx_factory(".csv")
        xlsx_path = tmp_xlsx_factory()
        Path(csv_path).write_text(f"v\n{text}\n", encoding="utf-8")
        xlsxturbo.csv_to_xlsx(csv_path, xlsx_path)
        cell = _cell(xlsx_path, "A2")
        assert not isinstance(cell, str), f"{text!r} became text"
        assert cell == expected

    def test_a_float_is_not_subject_to_the_integer_cutoff(self, tmp_xlsx: str) -> None:
        """The control: a float past 2^53 stays a number.

        The cutoff exists to stop an *integer* silently becoming a different
        integer. A float is already a double, so writing it as text would
        preserve nothing and cost the reader a numeric cell.

        Compared as a float on purpose. Excel stores the number in decimal and
        openpyxl hands back an ``int`` when those digits have no fractional
        part, so an exact ``int`` comparison would fail on a value that
        round-tripped perfectly.
        """
        frame = pd.DataFrame({"v": [1e17]})  # ~11x the integer cutoff
        xlsxturbo.df_to_xlsx(frame, tmp_xlsx)
        cell = _cell(tmp_xlsx, "A2")
        assert not isinstance(cell, str)
        assert float(cell) == 1e17


class TestDateSerialBoundary:
    """Excel's 1900 leap-year gap, through the CSV type-detection path."""

    @pytest.mark.parametrize(
        ("text", "stays_text"),
        [
            ("1899-12-29", True),  # before the epoch: negative serial
            ("1899-12-30", True),  # the epoch itself, serial 0
            ("1900-01-01", True),  # Excel's day 1, one off from the calendar
            ("1900-02-28", True),  # last day before the gap
            (FIRST_SANE_DATE, False),  # serial 61: the first day both agree on
            ("1900-03-02", False),
            ("2024-01-15", False),
            ("9999-12-31", False),  # Excel's last representable day
        ],
    )
    def test_dates_below_serial_61_are_written_as_text(
        self, tmp_xlsx_factory: Callable[..., str], text: str, stays_text: bool
    ) -> None:
        """A date Excel would render a day early is written as text instead."""
        csv_path = tmp_xlsx_factory(".csv")
        xlsx_path = tmp_xlsx_factory()
        Path(csv_path).write_text(f"d\n{text}\n", encoding="utf-8")
        xlsxturbo.csv_to_xlsx(csv_path, xlsx_path)
        cell = _cell(xlsx_path, "A2")
        if stays_text:
            assert cell == text
        else:
            # openpyxl resolves the serial and the number format back into a
            # date, which is the check that matters: it is an independent
            # reader agreeing that the cell really is a date.
            assert not isinstance(cell, str)
            assert cell.strftime("%Y-%m-%d") == text

    def test_the_first_sane_date_carries_serial_61(
        self, tmp_xlsx_factory: Callable[..., str]
    ) -> None:
        """Pin the serial itself, not only that a date came back.

        Reading through openpyxl's date conversion would agree with an epoch
        shifted by any whole number of days, so the raw number is read too.
        """
        csv_path = tmp_xlsx_factory(".csv")
        xlsx_path = tmp_xlsx_factory()
        Path(csv_path).write_text(f"d\n{FIRST_SANE_DATE}\n", encoding="utf-8")
        xlsxturbo.csv_to_xlsx(csv_path, xlsx_path)

        workbook = load_workbook(xlsx_path)
        try:
            raw = active_ws(workbook)["A2"].value
        finally:
            workbook.close()
        # openpyxl applies Excel's own epoch, which includes the phantom leap
        # day, so its round-trip is the authority on what serial was stored.
        from openpyxl.utils.datetime import to_excel

        assert to_excel(raw) == FIRST_SANE_SERIAL

    def test_a_real_leap_day_is_written_as_a_date(
        self, tmp_xlsx_factory: Callable[..., str]
    ) -> None:
        """The control for the 1900 cases: 2000 was a leap year."""
        csv_path = tmp_xlsx_factory(".csv")
        xlsx_path = tmp_xlsx_factory()
        Path(csv_path).write_text("d\n2000-02-29\n", encoding="utf-8")
        xlsxturbo.csv_to_xlsx(csv_path, xlsx_path)
        cell = _cell(xlsx_path, "A2")
        assert not isinstance(cell, str)
        assert cell.strftime("%Y-%m-%d") == "2000-02-29"


class TestCellReferenceBoundary:
    """Excel's grid limits, at the exact cell where they bite."""

    @pytest.mark.parametrize("ref", ["XFD1", "A1048576", "XFD1048576"])
    def test_the_last_addressable_cells_are_writable(
        self, tmp_xlsx: str, ref: str
    ) -> None:
        """The far corners of the grid are inside the limit, not outside it."""
        frame = pd.DataFrame({"v": [1]})
        xlsxturbo.df_to_xlsx(frame, tmp_xlsx, cells={ref: "corner"})
        assert _cell(tmp_xlsx, ref) == "corner"

    @pytest.mark.parametrize(
        ("ref", "fragment"),
        [
            ("XFE1", "exceeds Excel's maximum column"),
            ("A1048577", "exceeds Excel's allowed limits"),
        ],
    )
    def test_one_step_past_the_grid_is_refused(
        self, tmp_xlsx: str, ref: str, fragment: str
    ) -> None:
        """Both limits are enforced, and each says which one was hit.

        They are enforced in different places -- the column bound in this
        crate's own parser, the row bound by the Excel writer underneath -- so
        the messages differ. Both are ``ConfigurationError``, which is the part
        a caller depends on.
        """
        frame = pd.DataFrame({"v": [1]})
        with pytest.raises(xlsxturbo.ConfigurationError, match=fragment):
            xlsxturbo.df_to_xlsx(frame, tmp_xlsx, cells={ref: "over"})
