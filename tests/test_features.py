"""
Tests for xlsxturbo features (v0.6.0+):
- Global column width cap, table names, header styling (v0.6.0)
- Column formatting with wildcards (v0.7.0)
- Date order for CSV parsing, edge cases (v0.8.0)
- Formula columns, merged cells, hyperlinks (v0.9.0)
- Comments, validations, rich_text, images (v0.10.0)
"""

import xlsxturbo
import pandas as pd
import polars as pl
import tempfile
import os

# Only import openpyxl for verification if available
try:
    from openpyxl import load_workbook

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def get_temp_path():
    """Get a temporary file path that's closed for Windows compatibility"""
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    return path


class TestColumnWidthCap:
    """Tests for column_widths={'_all': value} feature"""

    def test_all_columns_capped(self):
        """'_all' key sets width for all columns"""
        df = pd.DataFrame({"A": ["x" * 100], "B": ["y" * 100], "C": ["z" * 100]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(df, path, column_widths={"_all": 20})
            assert os.path.exists(path)
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb.active
                for col in ["A", "B", "C"]:
                    assert ws.column_dimensions[col].width <= 21
                wb.close()
        finally:
            os.unlink(path)

    def test_specific_overrides_all(self):
        """Specific column width overrides '_all'"""
        df = pd.DataFrame({"A": ["x"], "B": ["y"], "C": ["z"]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(df, path, column_widths={0: 30, "_all": 10})
            assert os.path.exists(path)
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb.active
                assert ws.column_dimensions["A"].width > 25
                assert ws.column_dimensions["B"].width <= 11
                wb.close()
        finally:
            os.unlink(path)

    def test_all_with_autofit(self):
        """'_all' acts as cap when combined with autofit"""
        df = pd.DataFrame({"Short": ["x"], "VeryLongColumnName": ["y" * 100]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(df, path, autofit=True, column_widths={"_all": 25})
            assert os.path.exists(path)
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb.active
                # Long column should be capped at ~25
                assert ws.column_dimensions["B"].width <= 26
                wb.close()
        finally:
            os.unlink(path)

    def test_dfs_to_xlsx_per_sheet_all(self):
        """Per-sheet '_all' override in dfs_to_xlsx"""
        df1 = pd.DataFrame({"A": ["x" * 50]})
        df2 = pd.DataFrame({"B": ["y" * 50]})
        path = get_temp_path()
        try:
            xlsxturbo.dfs_to_xlsx(
                [
                    (df1, "Sheet1", {"column_widths": {"_all": 20}}),
                    (df2, "Sheet2", {"column_widths": {"_all": 40}}),
                ],
                path,
            )
            assert os.path.exists(path)
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                # Sheet1 should have narrower column than Sheet2
                w1 = wb["Sheet1"].column_dimensions["A"].width or 0
                w2 = wb["Sheet2"].column_dimensions["A"].width or 0
                assert w1 <= 21, f"Sheet1 col A width {w1} should be <= 21"
                assert w2 > 21, f"Sheet2 col A width {w2} should be > 21"
                wb.close()
        finally:
            os.unlink(path)


class TestTableName:
    """Tests for table_name parameter"""

    def test_explicit_table_name(self):
        """Explicit table name is applied"""
        df = pd.DataFrame({"A": [1, 2, 3]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(df, path, table_style="Medium2", table_name="MyTable")
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb.active
                tables = list(ws.tables.keys())
                assert "MyTable" in tables
                wb.close()
        finally:
            os.unlink(path)

    def test_table_name_sanitization(self):
        """Invalid characters are sanitized"""
        df = pd.DataFrame({"A": [1]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(
                df, path, table_style="Medium2", table_name="My Table-2024!"
            )
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb.active
                tables = list(ws.tables.keys())
                assert len(tables) == 1
                assert "_" in tables[0]  # Some characters replaced with underscore
                wb.close()
        finally:
            os.unlink(path)

    def test_table_name_starts_with_digit(self):
        """Table names starting with digit get underscore prefix"""
        df = pd.DataFrame({"A": [1]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(df, path, table_style="Medium2", table_name="123Data")
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb.active
                tables = list(ws.tables.keys())
                assert tables[0].startswith("_")
                wb.close()
        finally:
            os.unlink(path)

    def test_per_sheet_table_names(self):
        """Different table names per sheet"""
        df1 = pd.DataFrame({"A": [1]})
        df2 = pd.DataFrame({"B": [2]})
        path = get_temp_path()
        try:
            xlsxturbo.dfs_to_xlsx(
                [
                    (df1, "Sheet1", {"table_style": "Medium2", "table_name": "Table1"}),
                    (df2, "Sheet2", {"table_style": "Medium2", "table_name": "Table2"}),
                ],
                path,
            )
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                assert "Table1" in wb["Sheet1"].tables
                assert "Table2" in wb["Sheet2"].tables
                wb.close()
        finally:
            os.unlink(path)

    def test_no_table_name_without_table_style(self):
        """table_name is ignored if table_style is None"""
        df = pd.DataFrame({"A": [1]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(df, path, table_name="Ignored")
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb.active
                assert len(ws.tables) == 0
                wb.close()
        finally:
            os.unlink(path)


class TestHeaderFormat:
    """Tests for header_format parameter"""

    def test_bold_header(self):
        """Bold header is applied"""
        df = pd.DataFrame({"A": [1], "B": [2]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(df, path, header_format={"bold": True})
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb.active
                assert ws["A1"].font.bold == True
                assert ws["B1"].font.bold == True
                wb.close()
        finally:
            os.unlink(path)

    def test_header_background_color(self):
        """Background color is applied to header"""
        df = pd.DataFrame({"A": [1]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(df, path, header_format={"bg_color": "#4F81BD"})
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb.active
                # openpyxl uses ARGB format
                assert ws["A1"].fill.fgColor.rgb == "FF4F81BD"
                wb.close()
        finally:
            os.unlink(path)

    def test_header_font_color(self):
        """Font color is applied to header"""
        df = pd.DataFrame({"A": [1]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(df, path, header_format={"font_color": "white"})
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb.active
                assert ws["A1"].font.color.rgb == "FFFFFFFF"
                wb.close()
        finally:
            os.unlink(path)

    def test_combined_header_format(self):
        """Multiple format options combined"""
        df = pd.DataFrame({"A": [1]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(
                df,
                path,
                header_format={"bold": True, "bg_color": "#4F81BD", "font_color": "white"},
            )
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb.active
                assert ws["A1"].font.bold == True
                assert ws["A1"].fill.fgColor.rgb == "FF4F81BD"
                assert ws["A1"].font.color.rgb == "FFFFFFFF"
                wb.close()
        finally:
            os.unlink(path)

    def test_header_format_no_header(self):
        """header_format ignored when header=False"""
        df = pd.DataFrame({"A": [1]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(df, path, header=False, header_format={"bold": True})
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb.active
                # First row should be data, not header
                assert ws["A1"].value == 1
                wb.close()
        finally:
            os.unlink(path)

    def test_per_sheet_header_format(self):
        """Different header formats per sheet"""
        df1 = pd.DataFrame({"A": [1]})
        df2 = pd.DataFrame({"B": [2]})
        path = get_temp_path()
        try:
            xlsxturbo.dfs_to_xlsx(
                [
                    (df1, "Sheet1", {"header_format": {"bold": True}}),
                    (df2, "Sheet2", {"header_format": {"bg_color": "#FF0000"}}),
                ],
                path,
            )
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                assert wb["Sheet1"]["A1"].font.bold == True
                assert wb["Sheet2"]["A1"].fill.fgColor.rgb == "FFFF0000"
                wb.close()
        finally:
            os.unlink(path)


class TestBackwardCompatibility:
    """Ensure existing functionality still works"""

    def test_old_column_widths_still_works(self):
        """Integer key column_widths still works"""
        df = pd.DataFrame({"A": [1], "B": [2]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(df, path, column_widths={0: 20, 1: 30})
            assert os.path.exists(path)
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb.active
                assert ws.column_dimensions["A"].width > 15
                assert ws.column_dimensions["B"].width > 25
                wb.close()
        finally:
            os.unlink(path)

    def test_table_style_without_table_name(self):
        """table_style works without table_name (existing behavior)"""
        df = pd.DataFrame({"A": [1, 2, 3]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(df, path, table_style="Medium9")
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb.active
                assert len(ws.tables) == 1
                wb.close()
        finally:
            os.unlink(path)


class TestPolarsSupport:
    """Ensure all features work with polars DataFrames"""

    def test_polars_column_width_cap(self):
        df = pl.DataFrame({"A": ["x" * 100], "B": ["y" * 100]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(df, path, column_widths={"_all": 20})
            assert os.path.exists(path)
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb.active
                assert ws.column_dimensions["A"].width <= 21
                wb.close()
        finally:
            os.unlink(path)

    def test_polars_table_name(self):
        df = pl.DataFrame({"A": [1, 2, 3]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(df, path, table_style="Medium2", table_name="PolarsTable")
            assert os.path.exists(path)
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb.active
                assert "PolarsTable" in ws.tables
                wb.close()
        finally:
            os.unlink(path)

    def test_polars_header_format(self):
        df = pl.DataFrame({"A": [1], "B": [2]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(df, path, header_format={"bold": True})
            assert os.path.exists(path)
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb.active
                assert ws["A1"].font.bold == True
                wb.close()
        finally:
            os.unlink(path)


class TestAllFeaturesCombined:
    """Test using all new features together"""

    def test_all_features_df_to_xlsx(self):
        """All features work together in df_to_xlsx"""
        df = pd.DataFrame(
            {"Name": ["Alice", "Bob"], "Score": [95, 87], "Grade": ["A", "B"]}
        )
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(
                df,
                path,
                autofit=True,
                table_style="Medium2",
                table_name="StudentScores",
                column_widths={"_all": 30, 0: 20},
                header_format={"bold": True, "bg_color": "#4F81BD", "font_color": "white"},
                freeze_panes=True,
            )
            assert os.path.exists(path)
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb.active
                assert "StudentScores" in ws.tables
                # Note: table_style overrides header_format styling
                # This is expected Excel behavior - tables have their own header styles
                wb.close()
        finally:
            os.unlink(path)

    def test_all_features_dfs_to_xlsx(self):
        """All features work together in dfs_to_xlsx"""
        df1 = pd.DataFrame({"A": [1, 2], "B": [3, 4]})
        df2 = pd.DataFrame({"X": ["a", "b"], "Y": ["c", "d"]})
        path = get_temp_path()
        try:
            xlsxturbo.dfs_to_xlsx(
                [
                    (
                        df1,
                        "Numbers",
                        {
                            "table_style": "Medium2",
                            "table_name": "NumbersTable",
                            "header_format": {"bold": True},
                            "column_widths": {"_all": 15},
                        },
                    ),
                    (
                        df2,
                        "Letters",
                        {
                            "table_style": "Medium9",
                            "table_name": "LettersTable",
                            "header_format": {"bg_color": "#FF0000"},
                        },
                    ),
                ],
                path,
                autofit=True,
                freeze_panes=True,
            )
            assert os.path.exists(path)
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                assert "NumbersTable" in wb["Numbers"].tables
                assert "LettersTable" in wb["Letters"].tables
                wb.close()
        finally:
            os.unlink(path)


class TestEdgeCases:
    """Tests for edge cases and error handling"""

    def test_empty_dataframe(self):
        """Empty DataFrame writes successfully"""
        df = pd.DataFrame({"A": [], "B": []})
        path = get_temp_path()
        try:
            rows, cols = xlsxturbo.df_to_xlsx(df, path)
            assert rows == 1  # Just header
            assert cols == 2
            assert os.path.exists(path)
        finally:
            os.unlink(path)

    def test_empty_dataframe_with_table_style(self):
        """Empty DataFrame with table_style writes without creating table"""
        df = pd.DataFrame({"A": [], "B": []})
        path = get_temp_path()
        try:
            rows, cols = xlsxturbo.df_to_xlsx(df, path, table_style="Medium2")
            assert os.path.exists(path)
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb.active
                # No table should be created for empty DataFrame
                assert len(ws.tables) == 0
                wb.close()
        finally:
            os.unlink(path)

    def test_invalid_table_style_raises_error(self):
        """Invalid table_style raises ValueError"""
        df = pd.DataFrame({"A": [1, 2]})
        path = get_temp_path()
        try:
            try:
                xlsxturbo.df_to_xlsx(df, path, table_style="InvalidStyle")
                assert False, "Expected ValueError for invalid table_style"
            except ValueError as e:
                assert "Unknown table_style" in str(e)
                assert "InvalidStyle" in str(e)
        finally:
            if os.path.exists(path):
                os.unlink(path)

    def test_invalid_hex_color_raises_error(self):
        """Invalid hex color format raises ValueError"""
        df = pd.DataFrame({"A": [1]})
        path = get_temp_path()
        try:
            try:
                xlsxturbo.df_to_xlsx(df, path, header_format={"bg_color": "#FF"})
                assert False, "Expected ValueError for invalid hex color"
            except ValueError as e:
                assert "expected 6 characters" in str(e)
        finally:
            if os.path.exists(path):
                os.unlink(path)

    def test_column_formats_order_preserved(self):
        """Column format patterns are matched in order (first match wins)"""
        df = pd.DataFrame({"price_usd": [1.0], "price_eur": [2.0], "other": [3.0]})
        path = get_temp_path()
        try:
            # The more specific pattern should be listed first to take priority
            xlsxturbo.df_to_xlsx(
                df,
                path,
                column_formats={
                    "price_usd": {"bg_color": "#FF0000"},  # Specific - should match first
                    "price_*": {"bg_color": "#0000FF"},  # General - should match price_eur
                },
            )
            assert os.path.exists(path)
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb.active
                # price_usd should be red (specific match)
                assert ws["A2"].fill.fgColor.rgb == "FFFF0000"
                # price_eur should be blue (wildcard match)
                assert ws["B2"].fill.fgColor.rgb == "FF0000FF"
                # other should have no background
                wb.close()
        finally:
            os.unlink(path)

    def test_empty_dataframe_no_header(self):
        """Empty DataFrame with header=False"""
        df = pd.DataFrame({"A": [], "B": []})
        path = get_temp_path()
        try:
            rows, cols = xlsxturbo.df_to_xlsx(df, path, header=False)
            assert rows == 0
            assert cols == 2
            assert os.path.exists(path)
        finally:
            os.unlink(path)


class TestDateOrder:
    """Tests for date_order parameter in csv_to_xlsx"""

    def test_date_order_us_parses_mdy(self):
        """US date order parses 01-02-2024 as January 2"""
        import csv
        from datetime import datetime

        # Create CSV with ambiguous date
        csv_path = get_temp_path().replace(".xlsx", ".csv")
        xlsx_path = get_temp_path()
        try:
            with open(csv_path, "w", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(["date", "value"])
                writer.writerow(["01-02-2024", "100"])

            xlsxturbo.csv_to_xlsx(csv_path, xlsx_path, date_order="us")

            if HAS_OPENPYXL:
                wb = load_workbook(xlsx_path)
                ws = wb.active
                cell_value = ws["A2"].value
                # openpyxl returns datetime objects for Excel dates
                assert isinstance(cell_value, datetime), f"Expected datetime, got {type(cell_value)}"
                # US format: 01-02-2024 = January 2, 2024
                assert cell_value.month == 1, f"Expected January (1), got month {cell_value.month}"
                assert cell_value.day == 2, f"Expected day 2, got day {cell_value.day}"
                wb.close()
        finally:
            if os.path.exists(csv_path):
                os.unlink(csv_path)
            if os.path.exists(xlsx_path):
                os.unlink(xlsx_path)

    def test_date_order_eu_parses_dmy(self):
        """European date order parses 01-02-2024 as February 1"""
        import csv
        from datetime import datetime

        csv_path = get_temp_path().replace(".xlsx", ".csv")
        xlsx_path = get_temp_path()
        try:
            with open(csv_path, "w", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(["date", "value"])
                writer.writerow(["01-02-2024", "100"])

            xlsxturbo.csv_to_xlsx(csv_path, xlsx_path, date_order="eu")

            if HAS_OPENPYXL:
                wb = load_workbook(xlsx_path)
                ws = wb.active
                cell_value = ws["A2"].value
                assert isinstance(cell_value, datetime), f"Expected datetime, got {type(cell_value)}"
                # EU format: 01-02-2024 = February 1, 2024
                assert cell_value.month == 2, f"Expected February (2), got month {cell_value.month}"
                assert cell_value.day == 1, f"Expected day 1, got day {cell_value.day}"
                wb.close()
        finally:
            if os.path.exists(csv_path):
                os.unlink(csv_path)
            if os.path.exists(xlsx_path):
                os.unlink(xlsx_path)

    def test_date_order_produces_different_results(self):
        """US and EU date orders produce different Excel dates for ambiguous input"""
        import csv
        from datetime import datetime

        csv_path = get_temp_path().replace(".xlsx", ".csv")
        xlsx_us = get_temp_path()
        xlsx_eu = get_temp_path()
        try:
            with open(csv_path, "w", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(["date"])
                writer.writerow(["03-04-2024"])  # Mar 4 (US) vs Apr 3 (EU)

            xlsxturbo.csv_to_xlsx(csv_path, xlsx_us, date_order="us")
            xlsxturbo.csv_to_xlsx(csv_path, xlsx_eu, date_order="eu")

            if HAS_OPENPYXL:
                wb_us = load_workbook(xlsx_us)
                wb_eu = load_workbook(xlsx_eu)
                us_value = wb_us.active["A2"].value
                eu_value = wb_eu.active["A2"].value
                wb_us.close()
                wb_eu.close()

                # Values should be different dates
                assert us_value != eu_value, "US and EU should produce different dates"
                # US: March 4, EU: April 3 (30 days difference)
                diff = abs((us_value - eu_value).days)
                assert diff == 30, f"Expected 30 day difference, got {diff}"
        finally:
            if os.path.exists(csv_path):
                os.unlink(csv_path)
            if os.path.exists(xlsx_us):
                os.unlink(xlsx_us)
            if os.path.exists(xlsx_eu):
                os.unlink(xlsx_eu)

    def test_invalid_date_order_raises(self):
        """Invalid date_order raises ValueError"""
        import csv

        csv_path = get_temp_path().replace(".xlsx", ".csv")
        xlsx_path = get_temp_path()
        try:
            with open(csv_path, "w", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(["a"])
                writer.writerow(["1"])

            try:
                xlsxturbo.csv_to_xlsx(csv_path, xlsx_path, date_order="invalid")
                assert False, "Expected ValueError"
            except ValueError as e:
                assert "invalid" in str(e).lower()
        finally:
            if os.path.exists(csv_path):
                os.unlink(csv_path)
            if os.path.exists(xlsx_path):
                os.unlink(xlsx_path)


class TestFormulaColumns:
    """Tests for formula_columns feature (v0.9.0)"""

    def test_basic_formula(self):
        """Formula column appended after data columns"""
        df = pd.DataFrame({"price": [100, 200], "quantity": [5, 3]})
        path = get_temp_path()
        try:
            rows, cols = xlsxturbo.df_to_xlsx(
                df, path, formula_columns={"Total": "=A{row}*B{row}"}
            )
            assert cols == 3  # price, quantity, Total
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb.active
                # Header should be "Total"
                assert ws["C1"].value == "Total"
                # Data rows should have formulas (openpyxl shows them as =formula)
                assert ws["C2"].value == "=A2*B2"
                assert ws["C3"].value == "=A3*B3"
                wb.close()
        finally:
            os.unlink(path)

    def test_multiple_formula_columns(self):
        """Multiple formula columns in order"""
        df = pd.DataFrame({"price": [100], "qty": [5], "tax": [0.1]})
        path = get_temp_path()
        try:
            rows, cols = xlsxturbo.df_to_xlsx(
                df,
                path,
                formula_columns={
                    "Subtotal": "=A{row}*B{row}",
                    "TaxAmt": "=D{row}*C{row}",
                },
            )
            assert cols == 5  # price, qty, tax, Subtotal, TaxAmt
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb.active
                assert ws["D1"].value == "Subtotal"
                assert ws["E1"].value == "TaxAmt"
                assert ws["D2"].value == "=A2*B2"
                assert ws["E2"].value == "=D2*C2"
                wb.close()
        finally:
            os.unlink(path)

    def test_formula_row_placeholder(self):
        """The {row} placeholder is correctly replaced per row"""
        df = pd.DataFrame({"A": [10, 20, 30]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(
                df, path, formula_columns={"Double": "=A{row}*2"}
            )
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb.active
                assert ws["B2"].value == "=A2*2"
                assert ws["B3"].value == "=A3*2"
                assert ws["B4"].value == "=A4*2"
                wb.close()
        finally:
            os.unlink(path)

    def test_formula_with_dfs_to_xlsx(self):
        """Formula columns work in multi-sheet mode"""
        df = pd.DataFrame({"A": [1, 2]})
        path = get_temp_path()
        try:
            xlsxturbo.dfs_to_xlsx(
                [(df, "Sheet1", {"formula_columns": {"Sum": "=A{row}+10"}})],
                path,
            )
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb["Sheet1"]
                assert ws["B1"].value == "Sum"
                assert ws["B2"].value == "=A2+10"
                wb.close()
        finally:
            os.unlink(path)


class TestMergedRanges:
    """Tests for merged_ranges feature (v0.9.0)"""

    def test_simple_merge(self):
        """Merge a range with text"""
        df = pd.DataFrame({"A": [1, 2], "B": [3, 4], "C": [5, 6]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(
                df,
                path,
                merged_ranges=[("A1:C1", "Title Row")],
            )
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb.active
                # Cell A1 should contain the merge text
                assert ws["A1"].value == "Title Row"
                # The range should be merged
                merged = [str(m) for m in ws.merged_cells.ranges]
                assert "A1:C1" in merged
                wb.close()
        finally:
            os.unlink(path)

    def test_merge_with_format(self):
        """Merge a range with custom formatting"""
        df = pd.DataFrame({"A": [1], "B": [2]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(
                df,
                path,
                merged_ranges=[
                    ("A1:B1", "Styled Merge", {"bold": True, "bg_color": "#4F81BD"})
                ],
            )
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb.active
                assert ws["A1"].value == "Styled Merge"
                assert ws["A1"].font.bold is True
                merged = [str(m) for m in ws.merged_cells.ranges]
                assert "A1:B1" in merged
                wb.close()
        finally:
            os.unlink(path)

    def test_multiple_merges(self):
        """Multiple merged ranges in same sheet"""
        df = pd.DataFrame({"A": [1, 2, 3], "B": [4, 5, 6], "C": [7, 8, 9]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(
                df,
                path,
                merged_ranges=[
                    ("A1:C1", "Top Title"),
                    ("A5:C5", "Bottom Title"),
                ],
            )
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb.active
                merged = [str(m) for m in ws.merged_cells.ranges]
                assert "A1:C1" in merged
                assert "A5:C5" in merged
                wb.close()
        finally:
            os.unlink(path)

    def test_merge_with_dfs_to_xlsx(self):
        """Merged ranges work per-sheet"""
        df = pd.DataFrame({"A": [1], "B": [2]})
        path = get_temp_path()
        try:
            xlsxturbo.dfs_to_xlsx(
                [(df, "Sheet1", {"merged_ranges": [("A1:B1", "Per-Sheet Merge")]})],
                path,
            )
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb["Sheet1"]
                assert ws["A1"].value == "Per-Sheet Merge"
                merged = [str(m) for m in ws.merged_cells.ranges]
                assert "A1:B1" in merged
                wb.close()
        finally:
            os.unlink(path)


class TestHyperlinks:
    """Tests for hyperlinks feature (v0.9.0)"""

    def test_basic_hyperlink(self):
        """Hyperlink with URL and display text"""
        df = pd.DataFrame({"Name": ["Example"]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(
                df,
                path,
                hyperlinks=[("B2", "https://example.com", "Example Site")],
            )
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb.active
                assert ws["B2"].hyperlink is not None
                assert "example.com" in ws["B2"].hyperlink.target
                wb.close()
        finally:
            os.unlink(path)

    def test_hyperlink_without_display_text(self):
        """Hyperlink with URL only (no display text)"""
        df = pd.DataFrame({"A": [1]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(
                df,
                path,
                hyperlinks=[("A2", "https://example.com")],
            )
            assert os.path.exists(path)
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb.active
                assert ws["A2"].hyperlink is not None
                wb.close()
        finally:
            os.unlink(path)

    def test_multiple_hyperlinks(self):
        """Multiple hyperlinks in same sheet"""
        df = pd.DataFrame({"A": [1, 2, 3]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(
                df,
                path,
                hyperlinks=[
                    ("B1", "https://one.com", "One"),
                    ("B2", "https://two.com", "Two"),
                    ("B3", "https://three.com", "Three"),
                ],
            )
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb.active
                assert ws["B1"].hyperlink is not None
                assert ws["B2"].hyperlink is not None
                assert ws["B3"].hyperlink is not None
                wb.close()
        finally:
            os.unlink(path)

    def test_hyperlinks_with_dfs_to_xlsx(self):
        """Hyperlinks work per-sheet in multi-sheet mode"""
        df = pd.DataFrame({"A": [1]})
        path = get_temp_path()
        try:
            xlsxturbo.dfs_to_xlsx(
                [
                    (
                        df,
                        "Sheet1",
                        {"hyperlinks": [("B1", "https://example.com", "Link")]},
                    )
                ],
                path,
            )
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb["Sheet1"]
                assert ws["B1"].hyperlink is not None
                wb.close()
        finally:
            os.unlink(path)


class TestCsvConversion:
    """Tests for csv_to_xlsx function"""

    def test_basic_csv(self):
        """Basic CSV with header and data rows"""
        import csv

        csv_path = get_temp_path().replace(".xlsx", ".csv")
        xlsx_path = get_temp_path()
        try:
            with open(csv_path, "w", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(["name", "age", "score"])
                writer.writerow(["Alice", "30", "95.5"])
                writer.writerow(["Bob", "25", "88.0"])

            rows, cols = xlsxturbo.csv_to_xlsx(csv_path, xlsx_path)
            assert rows == 3
            assert cols == 3
            if HAS_OPENPYXL:
                wb = load_workbook(xlsx_path)
                ws = wb.active
                assert ws["A1"].value == "name"
                assert ws["B2"].value == 30  # should be detected as integer
                assert ws["C2"].value == 95.5  # should be detected as float
                wb.close()
        finally:
            if os.path.exists(csv_path):
                os.unlink(csv_path)
            if os.path.exists(xlsx_path):
                os.unlink(xlsx_path)

    def test_csv_type_detection(self):
        """CSV type detection: int, float, bool, date, string"""
        import csv

        csv_path = get_temp_path().replace(".xlsx", ".csv")
        xlsx_path = get_temp_path()
        try:
            with open(csv_path, "w", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(["int", "float", "bool", "date", "text"])
                writer.writerow(["42", "3.14", "true", "2024-01-15", "hello"])

            xlsxturbo.csv_to_xlsx(csv_path, xlsx_path)
            if HAS_OPENPYXL:
                wb = load_workbook(xlsx_path)
                ws = wb.active
                assert ws["A2"].value == 42
                assert abs(ws["B2"].value - 3.14) < 0.001
                assert ws["C2"].value is True
                # Date should be a datetime object in openpyxl
                from datetime import datetime

                assert isinstance(ws["D2"].value, datetime)
                assert ws["E2"].value == "hello"
                wb.close()
        finally:
            if os.path.exists(csv_path):
                os.unlink(csv_path)
            if os.path.exists(xlsx_path):
                os.unlink(xlsx_path)

    def test_csv_special_values(self):
        """CSV with NaN, Inf, empty cells"""
        import csv

        csv_path = get_temp_path().replace(".xlsx", ".csv")
        xlsx_path = get_temp_path()
        try:
            with open(csv_path, "w", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(["a", "b", "c"])
                writer.writerow(["NaN", "Inf", ""])
                writer.writerow(["nan", "-Inf", "   "])

            rows, cols = xlsxturbo.csv_to_xlsx(csv_path, xlsx_path)
            assert rows == 3
            if HAS_OPENPYXL:
                wb = load_workbook(xlsx_path)
                ws = wb.active
                # NaN/Inf/empty should become empty strings or None
                # (written as empty string in write_cell for CellValue::Empty)
                wb.close()
        finally:
            if os.path.exists(csv_path):
                os.unlink(csv_path)
            if os.path.exists(xlsx_path):
                os.unlink(xlsx_path)

    def test_csv_parallel(self):
        """CSV parallel mode produces same output"""
        import csv

        csv_path = get_temp_path().replace(".xlsx", ".csv")
        xlsx_seq = get_temp_path()
        xlsx_par = get_temp_path()
        try:
            with open(csv_path, "w", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(["num", "text"])
                for i in range(100):
                    writer.writerow([str(i), f"row_{i}"])

            rows_s, cols_s = xlsxturbo.csv_to_xlsx(csv_path, xlsx_seq, parallel=False)
            rows_p, cols_p = xlsxturbo.csv_to_xlsx(csv_path, xlsx_par, parallel=True)
            assert rows_s == rows_p
            assert cols_s == cols_p
            if HAS_OPENPYXL:
                wb_s = load_workbook(xlsx_seq)
                wb_p = load_workbook(xlsx_par)
                ws_s = wb_s.active
                ws_p = wb_p.active
                # Spot check some cells match
                for row in [1, 2, 50, 101]:
                    assert ws_s[f"A{row}"].value == ws_p[f"A{row}"].value
                    assert ws_s[f"B{row}"].value == ws_p[f"B{row}"].value
                wb_s.close()
                wb_p.close()
        finally:
            if os.path.exists(csv_path):
                os.unlink(csv_path)
            if os.path.exists(xlsx_seq):
                os.unlink(xlsx_seq)
            if os.path.exists(xlsx_par):
                os.unlink(xlsx_par)

    def test_csv_with_sheet_name(self):
        """CSV conversion with custom sheet name"""
        import csv

        csv_path = get_temp_path().replace(".xlsx", ".csv")
        xlsx_path = get_temp_path()
        try:
            with open(csv_path, "w", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(["a"])
                writer.writerow(["1"])

            xlsxturbo.csv_to_xlsx(csv_path, xlsx_path, sheet_name="MySheet")
            if HAS_OPENPYXL:
                wb = load_workbook(xlsx_path)
                assert "MySheet" in wb.sheetnames
                wb.close()
        finally:
            if os.path.exists(csv_path):
                os.unlink(csv_path)
            if os.path.exists(xlsx_path):
                os.unlink(xlsx_path)


class TestComments:
    """Tests for comments/notes feature (v0.10.0)"""

    def test_simple_comment(self):
        """Simple string comment"""
        df = pd.DataFrame({"A": [1, 2, 3]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(df, path, comments={"A1": "This is a header note"})
            assert os.path.exists(path)
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb.active
                # openpyxl stores comments in ws.comments
                assert ws["A1"].comment is not None
                assert "header note" in ws["A1"].comment.text
                wb.close()
        finally:
            os.unlink(path)

    def test_comment_with_author(self):
        """Comment with author"""
        df = pd.DataFrame({"A": [1]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(
                df, path, comments={"A2": {"text": "Data note", "author": "John"}}
            )
            assert os.path.exists(path)
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb.active
                comment = ws["A2"].comment
                assert comment is not None
                assert "Data note" in comment.text
                assert comment.author == "John"
                wb.close()
        finally:
            os.unlink(path)

    def test_multiple_comments(self):
        """Multiple comments on different cells"""
        df = pd.DataFrame({"A": [1, 2], "B": [3, 4]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(
                df, path, comments={"A1": "Column A", "B1": "Column B", "A2": "First value"}
            )
            assert os.path.exists(path)
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb.active
                assert ws["A1"].comment is not None
                assert ws["B1"].comment is not None
                assert ws["A2"].comment is not None
                assert "Column A" in ws["A1"].comment.text
                wb.close()
        finally:
            os.unlink(path)


class TestValidations:
    """Tests for data validation feature (v0.10.0)"""

    def test_list_validation(self):
        """Dropdown list validation"""
        df = pd.DataFrame({"Status": ["Open", "Closed"], "Value": [1, 2]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(
                df,
                path,
                validations={"Status": {"type": "list", "values": ["Open", "Closed", "Pending"]}},
            )
            assert os.path.exists(path)
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb.active
                # Check that data validation exists
                assert len(ws.data_validations.dataValidation) > 0
                wb.close()
        finally:
            os.unlink(path)

    def test_number_validation(self):
        """Whole number range validation"""
        df = pd.DataFrame({"Score": [85, 90]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(
                df,
                path,
                validations={"Score": {"type": "whole_number", "min": 0, "max": 100}},
            )
            assert os.path.exists(path)
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb.active
                assert len(ws.data_validations.dataValidation) > 0
                wb.close()
        finally:
            os.unlink(path)

    def test_validation_with_messages(self):
        """Validation with input and error messages"""
        df = pd.DataFrame({"Value": [50]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(
                df,
                path,
                validations={
                    "Value": {
                        "type": "decimal",
                        "min": 0,
                        "max": 100,
                        "input_title": "Enter Value",
                        "input_message": "Must be between 0 and 100",
                        "error_title": "Invalid",
                        "error_message": "Value out of range",
                    }
                },
            )
            assert os.path.exists(path)
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb.active
                assert len(ws.data_validations.dataValidation) > 0
                dv = ws.data_validations.dataValidation[0]
                assert dv.promptTitle == "Enter Value"
                assert dv.errorTitle == "Invalid"
                wb.close()
        finally:
            os.unlink(path)

    def test_validation_pattern_matching(self):
        """Validation with column pattern"""
        df = pd.DataFrame({"score_a": [80], "score_b": [90], "name": ["Test"]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(
                df, path, validations={"score_*": {"type": "whole_number", "min": 0, "max": 100}}
            )
            assert os.path.exists(path)
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb.active
                # Should have validations on the score columns
                assert len(ws.data_validations.dataValidation) > 0
                wb.close()
        finally:
            os.unlink(path)


class TestRichText:
    """Tests for rich text feature (v0.10.0)"""

    def test_rich_text_bold(self):
        """Rich text with bold formatting"""
        df = pd.DataFrame({"A": [1]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(
                df, path, rich_text={"A1": [("Bold", {"bold": True}), " normal"]}
            )
            assert os.path.exists(path)
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb.active
                # openpyxl may have issues reading rich text, just verify file exists
                wb.close()
        finally:
            os.unlink(path)

    def test_rich_text_mixed_formats(self):
        """Rich text with multiple format segments"""
        df = pd.DataFrame({"A": [1]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(
                df,
                path,
                rich_text={
                    "B2": [
                        ("Red text", {"font_color": "red"}),
                        " and ",
                        ("blue text", {"font_color": "blue", "italic": True}),
                    ]
                },
            )
            assert os.path.exists(path)
        finally:
            os.unlink(path)

    def test_rich_text_plain_segments(self):
        """Rich text with plain string segments"""
        df = pd.DataFrame({"A": [1]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(
                df, path, rich_text={"A3": ["Plain text ", ("bold", {"bold": True}), " more plain"]}
            )
            assert os.path.exists(path)
        finally:
            os.unlink(path)


class TestImages:
    """Tests for images feature (v0.10.0)"""

    def test_image_simple_path(self):
        """Image with simple path"""
        df = pd.DataFrame({"A": [1, 2, 3]})
        path = get_temp_path()
        # Create a small test image (1x1 white pixel PNG)
        import base64

        # Smallest valid PNG (1x1 white pixel)
        png_data = base64.b64decode(
            "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR42mP8z8BQDwAEhQGAhKmMIQAAAABJRU5ErkJggg=="
        )
        img_path = get_temp_path().replace(".xlsx", ".png")
        try:
            with open(img_path, "wb") as f:
                f.write(png_data)

            xlsxturbo.df_to_xlsx(df, path, images={"D1": img_path})
            assert os.path.exists(path)
        finally:
            if os.path.exists(path):
                os.unlink(path)
            if os.path.exists(img_path):
                os.unlink(img_path)

    def test_image_with_options(self):
        """Image with scaling options"""
        df = pd.DataFrame({"A": [1]})
        path = get_temp_path()
        import base64

        png_data = base64.b64decode(
            "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR42mP8z8BQDwAEhQGAhKmMIQAAAABJRU5ErkJggg=="
        )
        img_path = get_temp_path().replace(".xlsx", ".png")
        try:
            with open(img_path, "wb") as f:
                f.write(png_data)

            xlsxturbo.df_to_xlsx(
                df,
                path,
                images={"B5": {"path": img_path, "scale_width": 2.0, "scale_height": 2.0}},
            )
            assert os.path.exists(path)
        finally:
            if os.path.exists(path):
                os.unlink(path)
            if os.path.exists(img_path):
                os.unlink(img_path)


class TestV10AllFeatures:
    """Tests combining v0.10.0 features"""

    def test_all_new_features_together(self):
        """All v0.10.0 features work together"""
        df = pd.DataFrame({"Name": ["Alice", "Bob"], "Score": [85, 92]})
        path = get_temp_path()
        import base64

        png_data = base64.b64decode(
            "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR42mP8z8BQDwAEhQGAhKmMIQAAAABJRU5ErkJggg=="
        )
        img_path = get_temp_path().replace(".xlsx", ".png")
        try:
            with open(img_path, "wb") as f:
                f.write(png_data)

            xlsxturbo.df_to_xlsx(
                df,
                path,
                comments={"A1": "Names column", "B1": {"text": "Scores", "author": "Test"}},
                validations={"Score": {"type": "whole_number", "min": 0, "max": 100}},
                rich_text={"D1": [("Legend:", {"bold": True}), " student scores"]},
                images={"E1": img_path},
            )
            assert os.path.exists(path)
        finally:
            if os.path.exists(path):
                os.unlink(path)
            if os.path.exists(img_path):
                os.unlink(img_path)

    def test_new_features_with_dfs_to_xlsx(self):
        """New features work with dfs_to_xlsx"""
        df1 = pd.DataFrame({"A": [1, 2]})
        df2 = pd.DataFrame({"B": [3, 4]})
        path = get_temp_path()
        try:
            xlsxturbo.dfs_to_xlsx(
                [
                    (df1, "Sheet1", {"comments": {"A1": "First sheet header"}}),
                    (df2, "Sheet2", {"validations": {"B": {"type": "whole_number", "min": 0, "max": 10}}}),
                ],
                path,
            )
            assert os.path.exists(path)
        finally:
            os.unlink(path)


class TestErrorPaths:
    """Tests for error handling (v0.10.0)"""

    def test_nonexistent_image_file_raises_error(self):
        """Non-existent image file raises clear error"""
        df = pd.DataFrame({"A": [1]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(df, path, images={"B1": "/nonexistent/path/to/image.png"})
            assert False, "Should have raised an error"
        except ValueError as e:
            assert "Failed to load image" in str(e) or "image" in str(e).lower()
        finally:
            if os.path.exists(path):
                os.unlink(path)

    def test_validation_list_exceeds_255_chars_raises_error(self):
        """Validation list exceeding 255 chars raises clear error"""
        df = pd.DataFrame({"Status": ["A"]})
        path = get_temp_path()
        # Create values that exceed 255 chars total
        long_values = ["A" * 100, "B" * 100, "C" * 100]  # 300+ chars
        try:
            xlsxturbo.df_to_xlsx(
                df, path, validations={"Status": {"type": "list", "values": long_values}}
            )
            assert False, "Should have raised an error"
        except ValueError as e:
            assert "255" in str(e) and "character" in str(e).lower()
        finally:
            if os.path.exists(path):
                os.unlink(path)

    def test_invalid_validation_config_raises_error(self):
        """Invalid validation config (not a dict) raises clear error"""
        df = pd.DataFrame({"A": [1]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(df, path, validations={"A": "not_a_dict"})
            assert False, "Should have raised an error"
        except TypeError as e:
            assert "expected dict" in str(e).lower()
        finally:
            if os.path.exists(path):
                os.unlink(path)

    def test_invalid_rich_text_segment_raises_error(self):
        """Invalid rich_text segment (not string or tuple) raises clear error"""
        df = pd.DataFrame({"A": [1]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(df, path, rich_text={"A1": [123]})  # int is invalid
            assert False, "Should have raised an error"
        except TypeError as e:
            assert "segment" in str(e).lower() and ("string" in str(e).lower() or "tuple" in str(e).lower())
        finally:
            if os.path.exists(path):
                os.unlink(path)

    def test_wrong_type_column_widths_raises_error(self):
        """Passing a list instead of dict for column_widths raises TypeError"""
        df = pd.DataFrame({"A": [1]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(df, path, column_widths=[10, 20])
            assert False, "Should have raised TypeError"
        except TypeError as e:
            assert "expected dict" in str(e).lower()
            assert "column_widths" in str(e)
        finally:
            if os.path.exists(path):
                os.unlink(path)

    def test_wrong_type_header_format_raises_error(self):
        """Passing a string instead of dict for header_format raises TypeError"""
        df = pd.DataFrame({"A": [1]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(df, path, header_format="bold")
            assert False, "Should have raised TypeError"
        except TypeError as e:
            assert "expected dict" in str(e).lower()
            assert "header_format" in str(e)
        finally:
            if os.path.exists(path):
                os.unlink(path)

    def test_wrong_type_merged_ranges_raises_error(self):
        """Passing a dict instead of list for merged_ranges raises TypeError"""
        df = pd.DataFrame({"A": [1]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(df, path, merged_ranges={"A1:B1": "Title"})
            assert False, "Should have raised TypeError"
        except TypeError as e:
            assert "expected list" in str(e).lower()
            assert "merged_ranges" in str(e)
        finally:
            if os.path.exists(path):
                os.unlink(path)

    def test_wrong_type_hyperlinks_raises_error(self):
        """Passing a dict instead of list for hyperlinks raises TypeError"""
        df = pd.DataFrame({"A": [1]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(df, path, hyperlinks={"A1": "https://example.com"})
            assert False, "Should have raised TypeError"
        except TypeError as e:
            assert "expected list" in str(e).lower()
            assert "hyperlinks" in str(e)
        finally:
            if os.path.exists(path):
                os.unlink(path)

    def test_invalid_rich_text_not_list_raises_error(self):
        """Invalid rich_text value (not a list) raises clear error"""
        df = pd.DataFrame({"A": [1]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(df, path, rich_text={"A1": "not_a_list"})
            assert False, "Should have raised an error"
        except TypeError as e:
            assert "expected list" in str(e).lower()
        finally:
            if os.path.exists(path):
                os.unlink(path)


class TestConditionalFormatting:
    """Tests for conditional formatting feature (v0.8.0)"""

    def test_2_color_scale(self):
        """2-color scale conditional format"""
        df = pd.DataFrame({"Score": [10, 50, 90]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(
                df,
                path,
                conditional_formats={
                    "Score": {"type": "2_color_scale", "min_color": "#FF0000", "max_color": "#00FF00"}
                },
            )
            assert os.path.exists(path)
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb.active
                # openpyxl reads conditional formats
                assert len(ws.conditional_formatting) > 0
                wb.close()
        finally:
            os.unlink(path)

    def test_3_color_scale(self):
        """3-color scale conditional format"""
        df = pd.DataFrame({"Value": [1, 5, 10]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(
                df,
                path,
                conditional_formats={
                    "Value": {
                        "type": "3_color_scale",
                        "min_color": "#F8696B",
                        "mid_color": "#FFEB84",
                        "max_color": "#63BE7B",
                    }
                },
            )
            assert os.path.exists(path)
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb.active
                assert len(ws.conditional_formatting) > 0
                wb.close()
        finally:
            os.unlink(path)

    def test_data_bar(self):
        """Data bar conditional format"""
        df = pd.DataFrame({"Progress": [25, 50, 75, 100]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(
                df,
                path,
                conditional_formats={
                    "Progress": {"type": "data_bar", "bar_color": "#638EC6"}
                },
            )
            assert os.path.exists(path)
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb.active
                assert len(ws.conditional_formatting) > 0
                wb.close()
        finally:
            os.unlink(path)

    def test_icon_set(self):
        """Icon set conditional format"""
        df = pd.DataFrame({"Status": [1, 2, 3]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(
                df,
                path,
                conditional_formats={
                    "Status": {"type": "icon_set", "icon_type": "3_traffic_lights"}
                },
            )
            assert os.path.exists(path)
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb.active
                assert len(ws.conditional_formatting) > 0
                wb.close()
        finally:
            os.unlink(path)

    def test_conditional_format_with_pattern(self):
        """Conditional format with wildcard column pattern"""
        df = pd.DataFrame({"score_a": [80], "score_b": [60], "name": ["Alice"]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(
                df,
                path,
                conditional_formats={
                    "score_*": {"type": "2_color_scale", "min_color": "#FF0000", "max_color": "#00FF00"}
                },
            )
            assert os.path.exists(path)
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb.active
                # Should have conditional formats on both score columns
                assert len(ws.conditional_formatting) >= 1
                wb.close()
        finally:
            os.unlink(path)


class TestConstantMemoryMode:
    """Tests for constant_memory mode (v0.4.0)"""

    def test_basic_constant_memory(self):
        """File is created in constant memory mode"""
        df = pd.DataFrame({"A": list(range(100)), "B": list(range(100, 200))})
        path = get_temp_path()
        try:
            rows, cols = xlsxturbo.df_to_xlsx(df, path, constant_memory=True)
            assert rows > 0
            assert cols == 2
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb.active
                assert ws["A1"].value == "A"  # header
                assert ws["A2"].value == 0  # first data row
                assert ws["B2"].value == 100
                wb.close()
        finally:
            os.unlink(path)

    def test_constant_memory_silently_disables_features(self):
        """Features are silently disabled in constant memory mode (no crash)"""
        df = pd.DataFrame({"Score": [1, 2, 3]})
        path = get_temp_path()
        try:
            rows, cols = xlsxturbo.df_to_xlsx(
                df,
                path,
                constant_memory=True,
                # All these should be silently ignored:
                table_style="Medium9",
                freeze_panes=True,
                autofit=True,
                row_heights={0: 30},
                conditional_formats={"Score": {"type": "data_bar", "bar_color": "#638EC6"}},
                formula_columns={"Double": "=A{row}*2"},
                merged_ranges=[("A1:A1", "Merge")],
                hyperlinks=[("A1", "https://example.com", "Link")],
                comments={"A1": "Comment"},
                validations={"Score": {"type": "whole_number", "min": 0, "max": 100}},
                rich_text={"B1": [("Bold", {"bold": True})]},
            )
            assert rows > 0
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb.active
                # Table should NOT be created
                assert len(ws.tables) == 0
                # Data should still be written
                assert ws["A1"].value == "Score"
                assert ws["A2"].value == 1
                wb.close()
        finally:
            os.unlink(path)

    def test_constant_memory_with_column_widths(self):
        """Column widths still work in constant memory mode"""
        df = pd.DataFrame({"A": [1], "B": [2]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(df, path, constant_memory=True, column_widths={0: 25})
            assert os.path.exists(path)
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb.active
                assert ws.column_dimensions["A"].width > 20
                wb.close()
        finally:
            os.unlink(path)


class TestRowHeights:
    """Tests for row_heights parameter (v0.4.0)"""

    def test_basic_row_heights(self):
        """Set specific row heights"""
        df = pd.DataFrame({"A": [1, 2, 3]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(df, path, row_heights={0: 30, 2: 40})
            assert os.path.exists(path)
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb.active
                # openpyxl is 1-indexed; Excel may round heights slightly
                assert abs(ws.row_dimensions[1].height - 30) < 1
                assert abs(ws.row_dimensions[3].height - 40) < 1
                # Rows without explicit height should not have customHeight
                assert ws.row_dimensions[2].customHeight is False
                wb.close()
        finally:
            os.unlink(path)

    def test_row_heights_with_dfs_to_xlsx(self):
        """Row heights work per-sheet"""
        df = pd.DataFrame({"A": [1, 2]})
        path = get_temp_path()
        try:
            xlsxturbo.dfs_to_xlsx(
                [(df, "Sheet1", {"row_heights": {0: 25}})],
                path,
            )
            assert os.path.exists(path)
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb["Sheet1"]
                assert abs(ws.row_dimensions[1].height - 25) < 1
                wb.close()
        finally:
            os.unlink(path)

    def test_row_heights_ignored_in_constant_memory(self):
        """Row heights silently ignored in constant memory mode"""
        df = pd.DataFrame({"A": [1]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(df, path, constant_memory=True, row_heights={0: 50})
            assert os.path.exists(path)
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb.active
                # Row height should NOT be set (constant memory ignores it)
                # Default height is ~15, so it should not be 50
                assert ws.row_dimensions[1].height != 50 or ws.row_dimensions[1].height is None
                wb.close()
        finally:
            os.unlink(path)


if __name__ == "__main__":
    import sys

    # Run simple tests without pytest
    test_classes = [
        TestColumnWidthCap,
        TestTableName,
        TestHeaderFormat,
        TestBackwardCompatibility,
        TestPolarsSupport,
        TestAllFeaturesCombined,
        TestEdgeCases,
        TestDateOrder,
        TestFormulaColumns,
        TestMergedRanges,
        TestHyperlinks,
        TestCsvConversion,
        TestComments,
        TestValidations,
        TestRichText,
        TestImages,
        TestV10AllFeatures,
        TestErrorPaths,
        TestConditionalFormatting,
        TestConstantMemoryMode,
        TestRowHeights,
    ]

    failed = 0
    passed = 0

    for test_class in test_classes:
        instance = test_class()
        for method_name in dir(instance):
            if method_name.startswith("test_"):
                try:
                    getattr(instance, method_name)()
                    print(f"[PASS] {test_class.__name__}.{method_name}")
                    passed += 1
                except Exception as e:
                    print(f"[FAIL] {test_class.__name__}.{method_name}: {e}")
                    failed += 1

    print(f"\n{passed} passed, {failed} failed")
    sys.exit(1 if failed else 0)
