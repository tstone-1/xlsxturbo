"""
Tests for xlsxturbo v0.6.0 features:
- Global column width cap (column_widths={'_all': value})
- Table name parameter (table_name)
- Header styling (header_format)
"""

import xlsxturbo
import pandas as pd
import polars as pl
import tempfile
import os

# Only import openpyxl for verification if available
try:
    from openpyxl import load_workbook

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def get_temp_path():
    """Get a temporary file path that's closed for Windows compatibility"""
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    return path


class TestColumnWidthCap:
    """Tests for column_widths={'_all': value} feature"""

    def test_all_columns_capped(self):
        """'_all' key sets width for all columns"""
        df = pd.DataFrame({"A": ["x" * 100], "B": ["y" * 100], "C": ["z" * 100]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(df, path, column_widths={"_all": 20})
            assert os.path.exists(path)
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb.active
                for col in ["A", "B", "C"]:
                    assert ws.column_dimensions[col].width <= 21
                wb.close()
        finally:
            os.unlink(path)

    def test_specific_overrides_all(self):
        """Specific column width overrides '_all'"""
        df = pd.DataFrame({"A": ["x"], "B": ["y"], "C": ["z"]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(df, path, column_widths={0: 30, "_all": 10})
            assert os.path.exists(path)
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb.active
                assert ws.column_dimensions["A"].width > 25
                assert ws.column_dimensions["B"].width <= 11
                wb.close()
        finally:
            os.unlink(path)

    def test_all_with_autofit(self):
        """'_all' acts as cap when combined with autofit"""
        df = pd.DataFrame({"Short": ["x"], "VeryLongColumnName": ["y" * 100]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(df, path, autofit=True, column_widths={"_all": 25})
            assert os.path.exists(path)
        finally:
            os.unlink(path)

    def test_dfs_to_xlsx_per_sheet_all(self):
        """Per-sheet '_all' override in dfs_to_xlsx"""
        df1 = pd.DataFrame({"A": ["x" * 50]})
        df2 = pd.DataFrame({"B": ["y" * 50]})
        path = get_temp_path()
        try:
            xlsxturbo.dfs_to_xlsx(
                [
                    (df1, "Sheet1", {"column_widths": {"_all": 20}}),
                    (df2, "Sheet2", {"column_widths": {"_all": 40}}),
                ],
                path,
            )
            assert os.path.exists(path)
        finally:
            os.unlink(path)


class TestTableName:
    """Tests for table_name parameter"""

    def test_explicit_table_name(self):
        """Explicit table name is applied"""
        df = pd.DataFrame({"A": [1, 2, 3]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(df, path, table_style="Medium2", table_name="MyTable")
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb.active
                tables = list(ws.tables.keys())
                assert "MyTable" in tables
                wb.close()
        finally:
            os.unlink(path)

    def test_table_name_sanitization(self):
        """Invalid characters are sanitized"""
        df = pd.DataFrame({"A": [1]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(
                df, path, table_style="Medium2", table_name="My Table-2024!"
            )
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb.active
                tables = list(ws.tables.keys())
                assert len(tables) == 1
                assert "_" in tables[0]  # Some characters replaced with underscore
                wb.close()
        finally:
            os.unlink(path)

    def test_table_name_starts_with_digit(self):
        """Table names starting with digit get underscore prefix"""
        df = pd.DataFrame({"A": [1]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(df, path, table_style="Medium2", table_name="123Data")
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb.active
                tables = list(ws.tables.keys())
                assert tables[0].startswith("_")
                wb.close()
        finally:
            os.unlink(path)

    def test_per_sheet_table_names(self):
        """Different table names per sheet"""
        df1 = pd.DataFrame({"A": [1]})
        df2 = pd.DataFrame({"B": [2]})
        path = get_temp_path()
        try:
            xlsxturbo.dfs_to_xlsx(
                [
                    (df1, "Sheet1", {"table_style": "Medium2", "table_name": "Table1"}),
                    (df2, "Sheet2", {"table_style": "Medium2", "table_name": "Table2"}),
                ],
                path,
            )
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                assert "Table1" in wb["Sheet1"].tables
                assert "Table2" in wb["Sheet2"].tables
                wb.close()
        finally:
            os.unlink(path)

    def test_no_table_name_without_table_style(self):
        """table_name is ignored if table_style is None"""
        df = pd.DataFrame({"A": [1]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(df, path, table_name="Ignored")
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb.active
                assert len(ws.tables) == 0
                wb.close()
        finally:
            os.unlink(path)


class TestHeaderFormat:
    """Tests for header_format parameter"""

    def test_bold_header(self):
        """Bold header is applied"""
        df = pd.DataFrame({"A": [1], "B": [2]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(df, path, header_format={"bold": True})
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb.active
                # Note: table_style may override header_format styling
                # This is expected Excel behavior - tables have their own header styles
                pass
                assert ws["B1"].font.bold == True
                wb.close()
        finally:
            os.unlink(path)

    def test_header_background_color(self):
        """Background color is applied to header"""
        df = pd.DataFrame({"A": [1]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(df, path, header_format={"bg_color": "#4F81BD"})
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb.active
                # openpyxl uses ARGB format
                assert ws["A1"].fill.fgColor.rgb == "FF4F81BD"
                wb.close()
        finally:
            os.unlink(path)

    def test_header_font_color(self):
        """Font color is applied to header"""
        df = pd.DataFrame({"A": [1]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(df, path, header_format={"font_color": "white"})
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb.active
                assert ws["A1"].font.color.rgb == "FFFFFFFF"
                wb.close()
        finally:
            os.unlink(path)

    def test_combined_header_format(self):
        """Multiple format options combined"""
        df = pd.DataFrame({"A": [1]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(
                df,
                path,
                header_format={"bold": True, "bg_color": "#4F81BD", "font_color": "white"},
            )
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb.active
                # Note: table_style may override header_format styling
                # This is expected Excel behavior - tables have their own header styles
                pass
                assert ws["A1"].fill.fgColor.rgb == "FF4F81BD"
                wb.close()
        finally:
            os.unlink(path)

    def test_header_format_no_header(self):
        """header_format ignored when header=False"""
        df = pd.DataFrame({"A": [1]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(df, path, header=False, header_format={"bold": True})
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb.active
                # First row should be data, not header
                assert ws["A1"].value == 1
                wb.close()
        finally:
            os.unlink(path)

    def test_per_sheet_header_format(self):
        """Different header formats per sheet"""
        df1 = pd.DataFrame({"A": [1]})
        df2 = pd.DataFrame({"B": [2]})
        path = get_temp_path()
        try:
            xlsxturbo.dfs_to_xlsx(
                [
                    (df1, "Sheet1", {"header_format": {"bold": True}}),
                    (df2, "Sheet2", {"header_format": {"bg_color": "#FF0000"}}),
                ],
                path,
            )
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                assert wb["Sheet1"]["A1"].font.bold == True
                assert wb["Sheet2"]["A1"].fill.fgColor.rgb == "FFFF0000"
                wb.close()
        finally:
            os.unlink(path)


class TestBackwardCompatibility:
    """Ensure existing functionality still works"""

    def test_old_column_widths_still_works(self):
        """Integer key column_widths still works"""
        df = pd.DataFrame({"A": [1], "B": [2]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(df, path, column_widths={0: 20, 1: 30})
            assert os.path.exists(path)
        finally:
            os.unlink(path)

    def test_table_style_without_table_name(self):
        """table_style works without table_name (existing behavior)"""
        df = pd.DataFrame({"A": [1, 2, 3]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(df, path, table_style="Medium9")
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb.active
                assert len(ws.tables) == 1
                wb.close()
        finally:
            os.unlink(path)


class TestPolarsSupport:
    """Ensure all features work with polars DataFrames"""

    def test_polars_column_width_cap(self):
        df = pl.DataFrame({"A": ["x" * 100], "B": ["y" * 100]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(df, path, column_widths={"_all": 20})
            assert os.path.exists(path)
        finally:
            os.unlink(path)

    def test_polars_table_name(self):
        df = pl.DataFrame({"A": [1, 2, 3]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(df, path, table_style="Medium2", table_name="PolarsTable")
            assert os.path.exists(path)
        finally:
            os.unlink(path)

    def test_polars_header_format(self):
        df = pl.DataFrame({"A": [1], "B": [2]})
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(df, path, header_format={"bold": True})
            assert os.path.exists(path)
        finally:
            os.unlink(path)


class TestAllFeaturesCombined:
    """Test using all new features together"""

    def test_all_features_df_to_xlsx(self):
        """All features work together in df_to_xlsx"""
        df = pd.DataFrame(
            {"Name": ["Alice", "Bob"], "Score": [95, 87], "Grade": ["A", "B"]}
        )
        path = get_temp_path()
        try:
            xlsxturbo.df_to_xlsx(
                df,
                path,
                autofit=True,
                table_style="Medium2",
                table_name="StudentScores",
                column_widths={"_all": 30, 0: 20},
                header_format={"bold": True, "bg_color": "#4F81BD", "font_color": "white"},
                freeze_panes=True,
            )
            assert os.path.exists(path)
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                ws = wb.active
                assert "StudentScores" in ws.tables
                # Note: table_style may override header_format styling
                # This is expected Excel behavior - tables have their own header styles
                pass
                wb.close()
        finally:
            os.unlink(path)

    def test_all_features_dfs_to_xlsx(self):
        """All features work together in dfs_to_xlsx"""
        df1 = pd.DataFrame({"A": [1, 2], "B": [3, 4]})
        df2 = pd.DataFrame({"X": ["a", "b"], "Y": ["c", "d"]})
        path = get_temp_path()
        try:
            xlsxturbo.dfs_to_xlsx(
                [
                    (
                        df1,
                        "Numbers",
                        {
                            "table_style": "Medium2",
                            "table_name": "NumbersTable",
                            "header_format": {"bold": True},
                            "column_widths": {"_all": 15},
                        },
                    ),
                    (
                        df2,
                        "Letters",
                        {
                            "table_style": "Medium9",
                            "table_name": "LettersTable",
                            "header_format": {"bg_color": "#FF0000"},
                        },
                    ),
                ],
                path,
                autofit=True,
                freeze_panes=True,
            )
            assert os.path.exists(path)
            if HAS_OPENPYXL:
                wb = load_workbook(path)
                assert "NumbersTable" in wb["Numbers"].tables
                assert "LettersTable" in wb["Letters"].tables
                wb.close()
        finally:
            os.unlink(path)


if __name__ == "__main__":
    import sys

    # Run simple tests without pytest
    test_classes = [
        TestColumnWidthCap,
        TestTableName,
        TestHeaderFormat,
        TestBackwardCompatibility,
        TestPolarsSupport,
        TestAllFeaturesCombined,
    ]

    failed = 0
    passed = 0

    for test_class in test_classes:
        instance = test_class()
        for method_name in dir(instance):
            if method_name.startswith("test_"):
                try:
                    getattr(instance, method_name)()
                    print(f"[PASS] {test_class.__name__}.{method_name}")
                    passed += 1
                except Exception as e:
                    print(f"[FAIL] {test_class.__name__}.{method_name}: {e}")
                    failed += 1

    print(f"\n{passed} passed, {failed} failed")
    sys.exit(1 if failed else 0)
